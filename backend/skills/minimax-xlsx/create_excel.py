import json
import sys
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.path.insert(0, '/root/agent/backend/skills/minimax-xlsx')

# 读取API响应文件
json_data = open('/tmp/api_response.json').read()

data = json.loads(json_data)['data']

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "图像生成模型列表"

# 定义样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置表头
headers = ["模型ID", "名称", "描述", "上下文长度", "输入模态", "输出模态", "Tokenizer", 
           "提示价格($/token)", "完成价格($/token)", "支持参数", "创建时间", "知识截止日期"]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 填充数据
for row_idx, model in enumerate(data, 2):
    # 模型ID
    ws.cell(row=row_idx, column=1, value=model.get('id', '')).border = thin_border
    ws.cell(row=row_idx, column=1).alignment = cell_alignment
    
    # 名称
    ws.cell(row=row_idx, column=2, value=model.get('name', '')).border = thin_border
    ws.cell(row=row_idx, column=2).alignment = cell_alignment
    
    # 描述
    ws.cell(row=row_idx, column=3, value=model.get('description', '')).border = thin_border
    ws.cell(row=row_idx, column=3).alignment = cell_alignment
    
    # 上下文长度
    ws.cell(row=row_idx, column=4, value=model.get('context_length', '')).border = thin_border
    ws.cell(row=row_idx, column=4).alignment = cell_alignment
    
    # 输入模态
    input_mods = model.get('architecture', {}).get('input_modalities', [])
    ws.cell(row=row_idx, column=5, value=', '.join(input_mods) if input_mods else '').border = thin_border
    ws.cell(row=row_idx, column=5).alignment = cell_alignment
    
    # 输出模态
    output_mods = model.get('architecture', {}).get('output_modalities', [])
    ws.cell(row=row_idx, column=6, value=', '.join(output_mods) if output_mods else '').border = thin_border
    ws.cell(row=row_idx, column=6).alignment = cell_alignment
    
    # Tokenizer
    ws.cell(row=row_idx, column=7, value=model.get('architecture', {}).get('tokenizer', '')).border = thin_border
    ws.cell(row=row_idx, column=7).alignment = cell_alignment
    
    # 提示价格
    pricing = model.get('pricing', {})
    prompt_price = pricing.get('prompt', 'N/A')
    ws.cell(row=row_idx, column=8, value=prompt_price).border = thin_border
    ws.cell(row=row_idx, column=8).alignment = cell_alignment
    
    # 完成价格
    completion_price = pricing.get('completion', 'N/A')
    ws.cell(row=row_idx, column=9, value=completion_price).border = thin_border
    ws.cell(row=row_idx, column=9).alignment = cell_alignment
    
    # 支持参数
    params = model.get('supported_parameters', [])
    ws.cell(row=row_idx, column=10, value=', '.join(params) if params else '').border = thin_border
    ws.cell(row=row_idx, column=10).alignment = cell_alignment
    
    # 创建时间 (Unix时间戳转换)
    created = model.get('created')
    if created:
        created_date = datetime.datetime.fromtimestamp(created).strftime('%Y-%m-%d %H:%M:%S')
    else:
        created_date = ''
    ws.cell(row=row_idx, column=11, value=created_date).border = thin_border
    ws.cell(row=row_idx, column=11).alignment = cell_alignment
    
    # 知识截止日期
    ws.cell(row=row_idx, column=12, value=model.get('knowledge_cutoff', '')).border = thin_border
    ws.cell(row=row_idx, column=12).alignment = cell_alignment

# 调整列宽
column_widths = [35, 40, 60, 15, 20, 20, 12, 18, 18, 50, 20, 15]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + col_idx)].width = width

# 设置行高
ws.row_dimensions[1].height = 30
for row_idx in range(2, len(data) + 2):
    ws.row_dimensions[row_idx].height = 45

# 冻结首行
ws.freeze_panes = 'A2'

# 保存文件
output_path = '/root/agent/frontend/upload/openrouter_image_models.xlsx'
wb.save(output_path)
print(f"Excel文件已保存至: {output_path}")
